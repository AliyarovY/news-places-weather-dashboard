from openpyxl import load_workbook
from .models import Place


def import_places_from_xlsx(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name, latitude, longitude, rating = row
        place = Place.objects.create(
            name=name,
            latitude=latitude,
            longitude=longitude,
            rating=rating
        )
        place.save()
